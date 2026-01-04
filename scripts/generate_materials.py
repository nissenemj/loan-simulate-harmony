from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import os

# Ensure output directory exists
out_dir = "public/materials"
if not os.path.exists(out_dir):
    os.makedirs(out_dir)

# File paths
pdf_startti = os.path.join(out_dir, "Velkavapaus_Starttipaketti_v1.pdf")
pdf_checklist = os.path.join(out_dir, "Velkavapaus_Tarkistuslista_v1.pdf")
pdf_negotiate = os.path.join(out_dir, "Velkavapaus_Neuvottelupohjat_v1.pdf")
pdf_dashboard = os.path.join(out_dir, "Velkavapaus_KuukausiDashboard_v1.pdf")
xlsx_budget = os.path.join(out_dir, "Velkavapaus_Budjettipohja_v1.xlsx")
xlsx_snowball = os.path.join(out_dir, "Velkavapaus_Velkalumipallo_v1.xlsx")

# ---------- PDF Styles ----------
styles = getSampleStyleSheet()
body = ParagraphStyle("Body", parent=styles["BodyText"], fontName="Helvetica", fontSize=11, leading=15, spaceAfter=8)
h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=20, spaceAfter=12)
h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=14, spaceAfter=8)
h3 = ParagraphStyle("H3", parent=styles["Heading3"], fontName="Helvetica-Bold", fontSize=12, spaceAfter=6)
small = ParagraphStyle("Small", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=12, textColor=colors.grey)

# ========== 1. STARTTIPAKETTI PDF ==========
doc = SimpleDocTemplate(pdf_startti, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=16*mm, bottomMargin=16*mm)
story = []

story.append(Paragraph("Velkavapaus.fi - Starttipaketti v1", h1))
story.append(Paragraph(f"Paivays: {date.today().isoformat()}", small))
story.append(Spacer(1, 8))

intro = """
Tama on nopea, kaytannonlaheinen paketti sinulle, joka haluat saada velat hallintaan ilman itsesyytosta tai perfektionismia.
Tavoite: 7 paivan aikana selkeys - suunnitelma - ensimmainen konkreettinen askel.
"""
story.append(Paragraph(intro.strip(), body))

story.append(Paragraph("Mita tarvitset", h2))
need = """
* 30-60 minuuttia rauhallista aikaa (ja kyna).<br/>
* Paasy tiliotteisiin / luottokorttilaskuihin / lainaerittelyihin.<br/>
* Velkavapaus.fi:n Excel-pohjat: Budjettipohja v1 ja Velkalumipallo v1.
"""
story.append(Paragraph(need.strip(), body))

story.append(Paragraph("7 paivan mini-ohjelma", h2))
table_data = [
    ["Paiva", "Tehtava", "Kesto", "Tulos"],
    ["1", "Listaa kaikki velat ja pakolliset menot.", "30-60 min", "Velkalista"],
    ["2", "Rakennetaan budjetti: tulot, pakolliset, joustavat.", "30-45 min", "Budjettipohja taytetty"],
    ["3", "Etsi 1-3 pienta kuluvuotoa ja paata stoppi.", "20 min", "Ensimmainen saasto"],
    ["4", "Valitse strategia: lumipallo tai korkojahti.", "20-30 min", "Yksi valinta"],
    ["5", "Neuvottele: laskut/korot/erapaivat.", "15-30 min", "Ehto paranee"],
    ["6", "Aseta automaatio: veloille minimit + hyokkays.", "15 min", "Jatkuvuus"],
    ["7", "Tee seuraavan 30 paivan saanto.", "15 min", "Rytmi ja mittarit"],
]
tbl = Table(table_data, colWidths=[18*mm, 98*mm, 25*mm, 35*mm])
tbl.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("FONTSIZE", (0,0), (-1,0), 10),
    ("FONTSIZE", (0,1), (-1,-1), 9.5),
    ("VALIGN", (0,0), (-1,-1), "TOP"),
    ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
]))
story.append(tbl)
story.append(Spacer(1, 10))

story.append(Paragraph("Velkalumipallo kaytannossa", h2))
snowball = """
1) Maksa kaikista veloista minimierat ajallaan.<br/>
2) Valitse hyokkaysvelka: pienin saldo (lumipallo) tai korkein korko (korkojahti).<br/>
3) Laita kaikki ylimaarainen raha hyokkaysvelkaan.<br/>
4) Toista - maksusumma kasvaa kuin lumipallo.
"""
story.append(Paragraph(snowball.strip(), body))

story.append(PageBreak())

# LISAOSA 1: Psykologia
story.append(Paragraph("Velkaantumisen psykologia", h2))
psycho = """
Velka ei ole alykkyysongelma. Se on kuormitusongelma.<br/><br/>
Kun ihminen on stressaantunut, vasynyt tai epavarma tulevaisuudesta, aivot siirtyvat selviytymistilaan.
Tassa tilassa lyhyen aikavalin helpotus voittaa pitkan aikavalin jarjen.<br/><br/>
Tama ei ole heikkoutta. Tama on neurobiologiaa.<br/><br/>
Velkakierre ei katkea ryhdistautymalla, vaan vahentamalla kognitiivista kuormaa.
"""
story.append(Paragraph(psycho.strip(), body))

# LISAOSA 2: Strategiat
story.append(Paragraph("Velkastrategiat", h2))
strat = """
<b>1. Velkalumipallo (pienin saldo ensin)</b><br/>
Hyva jos: motivaatio on heikko, velkoja on monta, tarvitset nopeasti voittoja.<br/><br/>
<b>2. Korkojahti (korkein korko ensin)</b><br/>
Hyva jos: talous on jo hallinnassa, jaksat pitkaa projektia.<br/><br/>
<b>3. Hybridimalli</b><br/>
Maksa 1-2 pieninta velkaa pois ensin, sitten siirry korkojahtiin.
"""
story.append(Paragraph(strat.strip(), body))

doc.build(story)
print(f"Generated: {pdf_startti}")

# ========== 2. TARKISTUSLISTA PDF ==========
doc = SimpleDocTemplate(pdf_checklist, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=16*mm, bottomMargin=16*mm)
story = [
    Paragraph("Velkavapaus - 14 kohdan tarkistuslista", h1),
    Paragraph(f"Paivays: {date.today().isoformat()}", small),
    Spacer(1, 8)
]

items = [
    "Olen listannut kaikki velat (saldo, korko, erapaiva).",
    "Tiedan pakolliset kuukausimenoni.",
    "Budjetti on tehty (arvio + toteutunut).",
    "Minimierat on automatisoitu.",
    "Olen valinnut strategian.",
    "Hyokkayssumma on paatetty.",
    "Olen etsinyt kuluvuotoja.",
    "Yksi kuluvuoto on katkaistu.",
    "Olen ollut yhteydessa velkojaan.",
    "Erapaivat on synkronoitu.",
    "Minulla on viikkorutiini.",
    "Tiedan kuukauden nettosaldon.",
    "En ole ottanut uutta velkaa.",
    "Seuraava askel on kirjattu."
]

table = [[" ", "Tehtava"]] + [[" ", it] for it in items]
tbl = Table(table, colWidths=[12*mm, 150*mm])
tbl.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
    ("FONTSIZE", (0,0), (-1,-1), 10),
]))
story.append(tbl)
doc.build(story)
print(f"Generated: {pdf_checklist}")

# ========== 3. NEUVOTTELUPOHJAT PDF ==========
doc = SimpleDocTemplate(pdf_negotiate, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=16*mm, bottomMargin=16*mm)
story = [Paragraph("Velkavapaus - Neuvottelupohjat velkojille", h1)]

sections = [
    ("Maksueran pienennys", "Pyydan tilapaista maksueran pienennyst seuraavien 1-3 kuukauden ajaksi. Tavoitteeni on hoitaa velkani vastuullisesti."),
    ("Erapaivan siirto", "Onko mahdollista siirtaa erapaivaa palkkapaivani vastaavaksi?"),
    ("Korkojen tarkistus", "Onko lainani korkoon mahdollista tehda tarkistus moitteettoman maksuhistorian perusteella?"),
    ("Maksusuunnitelma", "Haluaisin sopia kirjallisen maksusuunnitelman.")
]

for title, text in sections:
    story.append(Paragraph(title, h2))
    story.append(Paragraph(text, body))
    story.append(Spacer(1, 8))

doc.build(story)
print(f"Generated: {pdf_negotiate}")

# ========== 4. KUUKAUSI-DASHBOARD PDF ==========
doc = SimpleDocTemplate(pdf_dashboard, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=16*mm, bottomMargin=16*mm)
story = [
    Paragraph("Velkavapaus - Kuukausi-dashboard", h1),
    Paragraph("Tayta kerran kuussa. Kolme lukua riittaa.", body)
]

dash = [
    ["Mittari", "Taman kuun arvo", "Edellinen kuukausi"],
    ["Velkojen kokonaismaara (EUR)", "", ""],
    ["Kuukauden nettosaldo (EUR)", "", ""],
    ["Hyokkayssumma (EUR)", "", ""],
]

tbl = Table(dash, colWidths=[70*mm, 40*mm, 40*mm])
tbl.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
    ("FONTSIZE", (0,0), (-1,-1), 11),
    ("TOPPADDING", (0,0), (-1,-1), 8),
    ("BOTTOMPADDING", (0,0), (-1,-1), 8),
]))
story.append(tbl)
story.append(Spacer(1, 16))
story.append(Paragraph("Muistiinpanot / paatokset seuraavalle kuulle:", h2))
notes_tbl = Table([[""]]*6, colWidths=[150*mm], rowHeights=[12*mm]*6)
notes_tbl.setStyle(TableStyle([("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey)]))
story.append(notes_tbl)
doc.build(story)
print(f"Generated: {pdf_dashboard}")

# ========== XLSX Styles ==========
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="F2F2F2")
bold = Font(bold=True)
money_fmt = u'#,##0.00 "EUR"'

# ========== 5. BUDJETTIPOHJA XLSX ==========
wb = Workbook()
ws = wb.active
ws.title = "Budjetti"

ws["A1"] = "Velkavapaus.fi - Budjettipohja v1"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:E1")

headers = ["Kategoria", "Arvio (EUR)", "Toteutunut (EUR)", "Ero (EUR)", "Huomio"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = bold
    cell.fill = hdr_fill
    cell.border = border

income_rows = ["Palkka", "Tuet", "Muut tulot"]
r = 4
for item in income_rows:
    ws.cell(row=r, column=1, value=item).border = border
    for c in range(2,5):
        ws.cell(row=r, column=c).number_format = money_fmt
        ws.cell(row=r, column=c).border = border
    ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    ws.cell(row=r, column=5).border = border
    r += 1

ws.cell(row=r, column=1, value="Yhteensa tulot").font = bold
ws.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").number_format = money_fmt
ws.cell(row=r, column=3, value=f"=SUM(C4:C{r-1})").number_format = money_fmt

for col in range(1, 6):
    ws.column_dimensions[get_column_letter(col)].width = 18

wb.save(xlsx_budget)
print(f"Generated: {xlsx_budget}")

# ========== 6. VELKALUMIPALLO XLSX ==========
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Velat"

ws2["A1"] = "Velkavapaus.fi - Velkalumipallo v1"
ws2["A1"].font = Font(bold=True, size=14)
ws2.merge_cells("A1:F1")

debt_headers = ["Velka", "Jaljella (EUR)", "Korko %", "Minimi/kk (EUR)", "Erapaiva", "Huomio"]
for i, h in enumerate(debt_headers, start=1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = bold
    cell.fill = hdr_fill
    cell.border = border

for rr in range(4, 16):
    for c in range(1, 7):
        ws2.cell(row=rr, column=c).border = border
    ws2.cell(row=rr, column=2).number_format = money_fmt
    ws2.cell(row=rr, column=4).number_format = money_fmt

ws2.cell(row=16, column=1, value="Yhteensa").font = bold
ws2.cell(row=16, column=2, value="=SUM(B4:B15)").number_format = money_fmt
ws2.cell(row=16, column=4, value="=SUM(D4:D15)").number_format = money_fmt

for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 18

wb2.save(xlsx_snowball)
print(f"Generated: {xlsx_snowball}")

print("\nAll materials generated successfully!")
